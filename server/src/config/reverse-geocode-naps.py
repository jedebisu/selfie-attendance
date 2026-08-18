#!/usr/bin/env python3
"""
Reverse geocode NAP coordinates using GADM boundaries.
Uses shapely for point-in-polygon tests.

Usage: python3 reverse-geocode-naps.py input.xlsx output.xlsx
"""

import json
import sys
import time
from shapely.geometry import shape, Point
from shapely.strtree import STRtree
import openpyxl


def load_gadm_boundaries(json_path):
    """Load GADM boundaries and build spatial index."""
    print(f'Loading boundaries from {json_path}...')
    with open(json_path, 'r') as f:
        data = json.load(f)
    
    features = data['features']
    print(f'  Loaded {len(features)} features')
    
    # Build geometries and index
    geometries = []
    properties = []
    for feat in features:
        geom = shape(feat['geometry'])
        props = feat['properties']
        geometries.append(geom)
        properties.append(props)
    
    # Build STRtree for fast spatial queries
    tree = STRtree(geometries)
    
    return tree, geometries, properties


def reverse_geocode(lat, lng, tree, geometries, properties):
    """Find province, city, barangay for a coordinate."""
    point = Point(lng, lat)  # Note: shapely uses (x, y) = (lng, lat)
    
    # Query the spatial index
    indices = tree.query(point)
    
    province = ''
    city = ''
    brgy = ''
    
    for idx in indices:
        if geometries[idx].contains(point):
            props = properties[idx]
            # Level 3 = barangay (has NAME_1, NAME_2, NAME_3)
            if 'NAME_3' in props and props.get('NAME_3'):
                brgy = props['NAME_3']
                city = props.get('NAME_2', '')
                province = props.get('NAME_1', '')
            # Level 2 = city/municipality (has NAME_1, NAME_2)
            elif 'NAME_2' in props and props.get('NAME_2'):
                city = props.get('NAME_2', '')
                province = props.get('NAME_1', '')
            break
    
    return province, city, brgy


def main():
    if len(sys.argv) < 3:
        print('Usage: python3 reverse-geocode-naps.py input.xlsx output.xlsx')
        sys.exit(1)
    
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    
    # Load boundaries
    tree_l2, geoms_l2, props_l2 = load_gadm_boundaries('/Users/yanggee/Downloads/gadm41_PHL_2.json')
    tree_l3, geoms_l3, props_l3 = load_gadm_boundaries('/Users/yanggee/Downloads/gadm41_PHL_3.json')
    
    # Read input Excel
    print(f'\nReading {input_path}...')
    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in.active
    
    # Get headers
    headers = [cell.value for cell in ws_in[1]]
    print(f'  Headers: {headers}')
    print(f'  Rows: {ws_in.max_row - 1}')
    
    # Create output workbook
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = 'NAP Facility Summary'
    
    # Add original headers + new columns
    out_headers = headers + ['PROVINCE_NAME', 'CITY_NAME', 'BRGY_NAME']
    ws_out.append(out_headers)
    
    # Process rows
    total = ws_in.max_row - 1
    geocoded = 0
    not_found = 0
    start_time = time.time()
    
    for i, row in enumerate(ws_in.iter_rows(min_row=2, values_only=True), start=1):
        lat = row[7]  # Latitude
        lng = row[8]  # Longitude
        
        province = ''
        city = ''
        brgy = ''
        
        if lat and lng:
            try:
                lat_f = float(lat)
                lng_f = float(lng)
                
                # Try barangay level first (level 3)
                province, city, brgy = reverse_geocode(lat_f, lng_f, tree_l3, geoms_l3, props_l3)
                
                # If no barangay found, try city level (level 2)
                if not brgy:
                    province, city, _ = reverse_geocode(lat_f, lng_f, tree_l2, geoms_l2, props_l2)
                
                if province or city:
                    geocoded += 1
                else:
                    not_found += 1
            except (ValueError, TypeError):
                not_found += 1
        
        # Write row with geocoded data
        out_row = list(row) + [province, city, brgy]
        ws_out.append(out_row)
        
        if i % 10000 == 0:
            elapsed = time.time() - start_time
            rate = i / elapsed if elapsed > 0 else 0
            eta = (total - i) / rate if rate > 0 else 0
            print(f'  Progress: {i}/{total} ({i*100//total}%) - {rate:.0f} rows/sec - ETA: {eta:.0f}s')
    
    # Save output
    print(f'\nSaving to {output_path}...')
    wb_out.save(output_path)
    
    elapsed = time.time() - start_time
    print(f'\nDone!')
    print(f'  Total: {total}')
    print(f'  Geocoded: {geocoded}')
    print(f'  Not found: {not_found}')
    print(f'  Time: {elapsed:.1f}s ({total/elapsed:.0f} rows/sec)')
    
    wb_in.close()


if __name__ == '__main__':
    main()
